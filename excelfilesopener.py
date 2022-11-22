import sys
from copy import copy

import openpyxl
from openpyxl.styles import Font
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import Qt, QSize
from PySide2.QtGui import QPixmap, QPainter, QIcon, QFont
from PySide2.QtWidgets import QHBoxLayout, QPushButton, QMainWindow, QApplication, QVBoxLayout, QWidget, QFileDialog, \
    QSpinBox, QLabel, QMessageBox, QListWidget, QComboBox, QToolBar, QAction, QAbstractButton
from PySide2.examples.widgets.itemviews.addressbook.tablemodel import TableModel


class PicButton(QAbstractButton):
    def __init__(self, pixmap, parent=None):
        super(PicButton, self).__init__(parent)
        self.pixmap = pixmap

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(event.rect(), self.pixmap)

    def sizeHint(self):
        return QSize(10, 50)


class TableModel(QtCore.QAbstractTableModel):
    def __init__(self, data):  # конструктор класса
        super().__init__()  # наследование от класса - родителя
        self._ws = data  # переменной self._ws передаём "выбранный лист" из нашей таблицы
        #self.filter = _filter  # передаём словарь, в котором указаны выставленные на панели
        # значения название шрифта, размер, курсив и жирность

    def flags(self, index):  # определение флагов для каждой ячейки в текущий момент
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def data(self, index, role):  # отображение наших "данных" в соответствии с определёнными параметрами

        if role == Qt.DisplayRole:  # отображение значения ячейки
            _value = self._ws[index.row() + 1][index.column()].value
            return _value
        elif role == Qt.FontRole:  # отображение данных, связанных со стилистикой текста (значения)
            # _fontName = self._ws[index.row() + 1][index.column()].font.name  # записывает название текущего шрифта
            # # данной ячейки
            # _fontBold = self._ws[index.row() + 1][index.column()].font.b  # возвращает флаг жирности: true - если жирный
            # _fontItalic = self._ws[index.row() + 1][index.column()].font.i  # возвращает флаг курсив: true - курсив
            # # _fontSize = self._ws[index.row() + 1][index.column()].font.size
            # font = QFont(_fontName, _fontItalic)  # создаём переменную класса Font c нашими параметрами
            #if _fontBold:
            #    font.setBold(True)  # устанавливаем параметр жирности в True, если текст в заданной ячейке жирный
            font = QFont()
            cell = self.getCell(index)
            font.setFamily(cell.font.name)
            font.setBold(cell.font.b)
            font.setItalic(cell.font.i)
            font.setPointSize(cell.font.sz)
            return font  # возвращаем переменную класса Font модели для правильного отображения

    def setData(self, index, value, role):  # функция, которая отвечает за изменение/установку данных
        if role == Qt.EditRole:  # если происходит изменение
            self._ws[index.row() + 1][index.column()].value = value  # изменяем значение ячейки по выбранным координатам
            # на листе Excel на введённое в программе
            return True
            # font = QFont()
            # font.setPointSize(self._ws[index.row() + 1][index.column()].font.sz)
            # font.setBold(self._ws[index.row() + 1][index.column()].font.b)
            # font.setItalic(self._ws[index.row() + 1][index.column()].font.i)

        return False

    def rowCount(self, index):  # возвращает количество строк в листе
        return self._ws.max_row

    def columnCount(self, index):  # возвращает количество столбцов в листе
        return self._ws.max_column

    def getCell(self, index):
        return self._ws.cell(index.row() + 1, index.column() + 1)

class MainWindow(QMainWindow):  # класс главного окна
    def __init__(self):  # конструктор главного окна
        super().__init__()  # наследование от класса - родителя (в нашем случае от QMainWindow)

        self.setWindowTitle("Excel Opener V Sashka")  # устанавливаем название окна
        self.setGeometry(300, 300, 800, 400)  # устанавливаем размеры начального окна

        # self.ui = Ui_MainWindow()
        # self.ui.setupUi(self)
        self.book = None  # инициализируем переменную, в которой будет храниться текущая таблица Excel
        self.sheet = None  # инициализируем переменную, в которой будет храниться текущий лист таблицы Excel
        self.file_name = ""
        # self.filter = {"name": "", "size": 0, "bold": False, "italic": False}  инициализируем массив filter
        # для последующего изменения параметров (
        self.mainToolbar = QToolBar("Main toolbar")  # создаём тулбар (верхнее меню)
        self.addToolBar(self.mainToolbar)  # связываем тулбар с главным окном

        self.open_action = QAction("Open")  # создаём кнопку "Open"
        self.open_action.setStatusTip("Open from your pc")  # устанавливаем подсказку "Open from your pc"
        self.open_action.triggered.connect(self.open_button_clicked)  # передаём управление функции при нажатии
        self.open_action.setShortcut(Qt.CTRL + Qt.Key_O)  # добавили быструю клавишу

        self.save_action = QAction("Save")  # создаём кнопку "Save"
        self.save_action.setStatusTip("Save document")  # устанавливаем подсказку
        self.save_action.triggered.connect(self.save_button_clicked)  # передаём управление функции при нажатии
        self.save_action.setShortcut(Qt.CTRL + Qt.Key_S)  # добавили бычтрую клавишу дл сохранения

        self.close_action = QAction("Close")  # создаём кнопку "Close"
        self.close_action.setStatusTip("Close current document")  # устанавливаем подсказку
        self.close_action.triggered.connect(self.close_button_clicked)  # передаём управление функции при нажатии

        self.filter_action = QAction("Help")  # создаём кнопку "Help"
        self.filter_action.setStatusTip("You can find some information about the program")  # устанавливаем подсказку
        self.filter_action.triggered.connect(self.filter_button_clicked)  # передаём управление функции при нажатии

        self.mainToolbar.addAction(self.open_action)  # добавляем кнопку (событие) в тулбар
        self.mainToolbar.addSeparator()  # добавляем разделитель
        self.mainToolbar.addAction(self.save_action)  # добавляем кнопку (событие) в тулбар
        self.mainToolbar.addSeparator()  # добавляем разделитель
        self.mainToolbar.addAction(self.close_action)  # добавляем кнопку (событие) в тулбар
        self.mainToolbar.addSeparator()  # добавляем разделитель
        self.mainToolbar.addAction(self.filter_action)  # добавляем кнопку (событие) в тулбар

        self.sheetCount = QVBoxLayout()  # создаём виджет, позволяющий располагать другие виджеты по вертикали
        self.sheetCountLabel = QLabel("Worksheet")  # создаём виджет "Текст" с названием "Worksheet"
        self.sheetListBox = QComboBox()  # создаём виджет, в котором можно выбрать лист из таблицы Excel
        self.sheetListBox.currentTextChanged.connect(self.sheetListBox_text_changed)  # передаём управление функции
        # при изменении значения отображаемого поля в виджете
        self.sheetCount.addWidget(self.sheetCountLabel)  # добавляем в вертикальный виджет название и выборку
        self.sheetCount.addWidget(self.sheetListBox)

        self.fontChanger = QComboBox()  # выбор шрифта
        self.fontChanger.addItems(["Calibri", "Arial", "Times New Roman"])  # добавляем начальные названия шрифтов
        self.fontChanger.currentTextChanged.connect(self.font_change)  # передаём управление об изменении шрифта через
        # виджет

        self.fontSize = QSpinBox()  # выбор размера шрифта
        self.fontSize.setMinimum(6)  # устанавливаем минимальное значение размера шрифта
        self.fontSize.setMaximum(20)  # устанавливаем максимальное значение размера шрифта
        self.fontSize.valueChanged.connect(self.fontSizeChanged)
        # self.boldLetter = PicButton(QPixmap("boldLetter.png"))
        self.boldLetter = QPushButton()  # создаём кнопку, которая будет изменять жирность шрифта
        self.boldLetter.setIcon(QIcon("boldLetter.png"))  # устаналиваем иконку на кнопку
        self.boldLetter.clicked.connect(self.bold_changes)

        self.italicLetter = QPushButton()  # создаём кнопку, которая будет изменять курсив шрифта
        self.italicLetter.setIcon(QIcon("cursevLetter.png"))  # устаналиваем иконку на кнопку
        self.italicLetter.clicked.connect(self.italic_changes)

        self.hLayout = QHBoxLayout()  # создаём виджет, позволяющий располагать другие виджеты по горизонтали
        self.hLayout.addWidget(self.fontChanger)  # добаляем виджет
        self.hLayout.addWidget(self.fontSize)  # добаляем виджет
        self.hLayout.addWidget(self.boldLetter)  # добаляем виджет
        self.hLayout.addWidget(self.italicLetter)  # добаляем виджет
        self.hLayout.addLayout(self.sheetCount)  # добаляем виджет

        self.vLayout = QVBoxLayout()  # создаём головной виджет вертикального расположения
        self.vLayout.addLayout(self.hLayout)  # добавляем туда наше "меню" для редактирования

        self.table = QtWidgets.QTableView()  # создаём виджет-отображение таблицы
        self.vLayout.addWidget(self.table)  # добавляем в головной виджет

        self.widget = QWidget()  # создаём виджет для отображения
        self.widget.setLayout(self.vLayout)  # вставляем в наш виджет, головной виджет вертикального расположения
        self.setCentralWidget(self.widget)  # выставляем виджет на передний план

    def open_button_clicked(self):  # нажатие кнопки "Open"
        current_file_name = QFileDialog.getOpenFileName(self)  # получаем название файла после выборки
        if current_file_name[0] != "":
            if self.book is not None:  # если у нас открыта таблица в Excel0
                self.book.close()  # то закрываем её
            self.sheet = None  # "удаляем" текущий лист
            self.sheetListBox.clear()  # очищаем виджет с названием листов таблицы
            self.table.setModel(None)  # "удаляем" текущую модель
            self.file_name = current_file_name[0]
            if self.file_name[-5:] == ".xlsx" or self.file_name[-4:] == ".xls":  # проверяем расширение файла
                self.book = openpyxl.load_workbook(self.file_name)  # загружаем текущую таблицу
                self.sheetListBox.addItems(self.book.sheetnames)  # добавляем названия листов таблицы

    def close_button_clicked(self):  # нажатие кнопки "Close"
        qMessage = QMessageBox()  # создаём диалоговой окно с вопросов: "Вы уверены, что хотите закрыть файл?"
        qMessage.setWindowTitle("Question")
        qMessage.setText("Are you sure you want to close the file?")
        qMessage.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        qMessage.setIcon(QMessageBox.Question)

        if self.book is not None:
            button = qMessage.exec_()

            if button == QMessageBox.Yes:  # проверка ответа пользователя, если да - обнуление всех данных
                self.book.close()
                self.sheet = None
                self.sheetListBox.clear()
                self.table.setModel(None)

    def save_button_clicked(self):  # нажатие кнопки "Save"
        try:
            self.book.save(self.file_name)  # попытка сохранить книгу
        except Exception:  # если выбрасывается исключение, то создаём уведомление об ошибке
            falseMessage = QMessageBox.critical(self, "Error", "Unreal to save this file")
        else:
            okMessage = QMessageBox()  # иначе создаём окно об успешном
            okMessage.setWindowTitle("Ok")  #
            okMessage.setText("The Excel file was saved successfully")
            okMessage.setIcon(QMessageBox.Information)
            okMessage.exec_()

    def filter_button_clicked(self):
        self.w = QBoxWindow()
        self.w.show()
        # self.table.model().filter =

    def sheetListBox_text_changed(self, str):
        try:
            self.sheet = self.book[str]
            self.model = TableModel(self.sheet)
            self.table.setModel(self.model)
        except KeyError:
            o = 1

    def font_change(self, s):
        # self.filter["name"] = s
        indexes = self.table.selectedIndexes()
        for index in indexes:
            cell = self.model.getCell(index)
            newFont = copy(cell.font)
            newFont.name = s
            cell.font = newFont
            self.model.dataChanged.emit(index, index)

    def bold_changes(self):
        indexes = self.table.selectedIndexes()
        for index in indexes:
            cell = self.model.getCell(index)
            newFont = copy(cell.font)
            newFont.b = not cell.font.b
            cell.font = newFont
            self.model.dataChanged.emit(index, index)

    def italic_changes(self):
        indexes = self.table.selectedIndexes()
        for index in indexes:
            cell = self.model.getCell(index)
            newFont = copy(cell.font)
            newFont.i = not cell.font.i
            cell.font = newFont
            self.model.dataChanged.emit(index, index)

    def fontSizeChanged(self, i):
        indexes = self.table.selectedIndexes()
        for index in indexes:
            cell = self.model.getCell(index)
            newFont = copy(cell.font)
            newFont.sz = i
            cell.font = newFont
            self.model.dataChanged.emit(index, index)

class QBoxWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Help")
        self.setGeometry(500, 400, 300, 300)

        self.shortcutsLabel = QLabel("Shortcuts")
        self.shortcutsLabel.setFont(QtGui.QFont("Calibri", 20))
        self.shortcutsLabel.setAlignment(Qt.AlignCenter)

        self.openLabel = QLabel("\"CTRL + O\" - open a file from the directory")
        self.openLabel.setFont(QtGui.QFont("Calibri", 10))
        self.openLabel.setAlignment(Qt.AlignCenter)

        self.saveLabel = QLabel("\"CTRL + S\" - save a file")
        self.saveLabel.setFont(QtGui.QFont("Calibri", 10))
        self.saveLabel.setAlignment(Qt.AlignCenter)

        self.vLayout = QVBoxLayout()
        self.vLayout.addWidget(self.shortcutsLabel)
        self.vLayout.addWidget(self.openLabel)
        self.vLayout.addWidget(self.saveLabel)

        self.widget = QWidget()
        self.widget.setLayout(self.vLayout)
        self.setCentralWidget(self.widget)


app = QApplication(sys.argv)
window = MainWindow()
window.show()
app.exec_()
